import argparse
import os
import re
import unicodedata
from datetime import date, datetime, timedelta
from urllib.parse import urlparse

import psycopg2
from openpyxl import load_workbook


MONTHS = {
    "janvier": 1,
    "fevrier": 2,
    "mars": 3,
    "avril": 4,
    "mai": 5,
    "juin": 6,
    "juillet": 7,
    "aout": 8,
    "septembre": 9,
    "octobre": 10,
    "novembre": 11,
    "decembre": 12,
}

MONTH_PATTERN = (
    r"janvier|f[ée]vrier|mars|avril|mai|juin|juillet|ao[uû]t|septembre|octobre|novembre|d[ée]cembre"
)
FRENCH_TEXT_DATE_RE = re.compile(
    rf"\b(?P<days>\d{{1,2}}(?:\s*(?:-|,|et)\s*\d{{1,2}})*)\s+"
    rf"(?P<month>{MONTH_PATTERN})"
    rf"(?:\s+(?P<year>\d{{4}}))?\b",
    flags=re.IGNORECASE,
)
SLASH_DATE_RE = re.compile(r"\b(?P<day>\d{1,2})/(?P<month>\d{1,2})/(?P<year>\d{4})\b")
WEEKDAY_TOKENS = {
    "lu",
    "lun",
    "ma",
    "mar",
    "me",
    "mer",
    "je",
    "jeu",
    "ve",
    "ven",
    "sa",
    "sam",
    "di",
    "dim",
}
MONTH_BLOCK_STARTS = tuple(range(1, 40, 4))
GRID_START_ROW = 5
GRID_END_ROW = 35
FOOTER_START_ROW = 36


def _normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", normalized).strip()


def _clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def _guess_type(label: str) -> str:
    normalized = _normalize_text(label).lower()
    if not normalized:
        return "evenement"
    if "vacances" in normalized:
        return "vacances"
    if "examen" in normalized or re.search(r"\bds\b", normalized):
        return "examen"
    if "revision" in normalized or "ratt" in normalized:
        return "revision"
    if normalized in {"p1", "p2"} or normalized.startswith("s1 p") or normalized.startswith("s2 p"):
        return "periode"
    if any(token in normalized for token in ("fete", "aid", "jour", "journee", "mawled")):
        return "jour_ferie"
    return "evenement"


def _build_merged_value_map(worksheet):
    merged_values = {}
    for merged_range in worksheet.merged_cells.ranges:
        anchor_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = anchor_value
    return merged_values


def _cell_value(worksheet, merged_values, row: int, col: int):
    value = worksheet.cell(row, col).value
    if value is None:
        value = merged_values.get((row, col))
    return value


def _extract_year_markers(worksheet, merged_values):
    markers = []
    for col in range(1, worksheet.max_column + 1):
        value = _cell_value(worksheet, merged_values, 3, col)
        text = _clean_text(value)
        if re.fullmatch(r"20\d{2}", text):
            markers.append((col, int(text)))
    return markers


def _year_for_column(year_markers, col: int) -> int | None:
    year = None
    for marker_col, marker_year in year_markers:
        if marker_col <= col:
            year = marker_year
        else:
            break
    return year


def _month_num_for_col(worksheet, merged_values, col: int) -> int | None:
    month_text = _normalize_text(_cell_value(worksheet, merged_values, 4, col)).lower()
    return MONTHS.get(month_text)


def _is_weekday_or_number(text: str) -> bool:
    normalized = _normalize_text(text).lower()
    return not normalized or normalized.isdigit() or normalized in WEEKDAY_TOKENS


def _pick_block_label(worksheet, merged_values, row: int, start_col: int) -> str:
    candidates = []
    for col in (start_col + 2, start_col + 3):
        text = _clean_text(_cell_value(worksheet, merged_values, row, col))
        if _is_weekday_or_number(text):
            continue
        candidates.append(text)

    if not candidates:
        return ""

    unique_candidates = []
    seen = set()
    for candidate in candidates:
        key = _normalize_text(candidate).lower()
        if key not in seen:
            unique_candidates.append(candidate)
            seen.add(key)

    return max(unique_candidates, key=len)


def _should_skip_grid_label(label: str) -> bool:
    normalized = _normalize_text(label).lower()
    if not normalized or normalized in WEEKDAY_TOKENS:
        return True
    return _guess_type(label) in {"jour_ferie", "periode"}


def _format_period_label(label: str, event_date: date) -> str:
    normalized = _normalize_text(label).upper()
    if normalized not in {"P1", "P2"}:
        return label
    semester = 1 if event_date.month >= 9 else 2
    return f"S{semester} {normalized}"


def _extract_week_markers_from_sheet(worksheet, merged_values, year_markers):
    markers = []

    for start_col in MONTH_BLOCK_STARTS:
        month_num = _month_num_for_col(worksheet, merged_values, start_col)
        year = _year_for_column(year_markers, start_col)
        if month_num is None or year is None:
            continue

        for row in range(GRID_START_ROW, GRID_END_ROW + 1):
            day_text = _clean_text(_cell_value(worksheet, merged_values, row, start_col))
            week_text = _clean_text(_cell_value(worksheet, merged_values, row, start_col + 3))
            if not day_text.isdigit() or not week_text.isdigit():
                continue

            try:
                marker_date = datetime(year, month_num, int(day_text)).date()
            except ValueError:
                continue

            markers.append((int(week_text), marker_date))

    markers.sort(key=lambda item: item[1])
    return markers


def _extract_period_events_from_sheet(worksheet):
    merged_values = _build_merged_value_map(worksheet)
    year_markers = _extract_year_markers(worksheet, merged_values)
    week_markers = _extract_week_markers_from_sheet(worksheet, merged_values, year_markers)
    events = []

    semester_index = 0
    for week_number, start_date in week_markers:
        if week_number == 1:
            semester_index += 1
            if semester_index > 2:
                continue
            events.append(
                {
                    "nom": f"S{semester_index} P1",
                    "date_debut": start_date,
                    "date_fin": start_date + timedelta(days=48),
                    "type": "periode",
                }
            )
        elif week_number == 8 and 1 <= semester_index <= 2:
            events.append(
                {
                    "nom": f"S{semester_index} P2",
                    "date_debut": start_date,
                    "date_fin": start_date + timedelta(days=48),
                    "type": "periode",
                }
            )

    return events


def _append_range(events, label: str, start_date: date, end_date: date):
    clean_label = re.sub(r"\s+", " ", label).strip()
    if not clean_label:
        return
    events.append(
        {
            "nom": clean_label,
            "date_debut": start_date,
            "date_fin": end_date,
            "type": _guess_type(clean_label),
        }
    )


def _merge_adjacent_ranges(events):
    merged_events = []
    for event in sorted(events, key=lambda item: (item["nom"].lower(), item["type"], item["date_debut"])):
        if not merged_events:
            merged_events.append(dict(event))
            continue

        previous = merged_events[-1]
        if (
            previous["nom"].lower() == event["nom"].lower()
            and previous["type"] == event["type"]
            and event["date_debut"] <= previous["date_fin"] + timedelta(days=1)
        ):
            previous["date_fin"] = max(previous["date_fin"], event["date_fin"])
            continue

        merged_events.append(dict(event))

    return merged_events


def _extract_grid_events_from_sheet(worksheet):
    merged_values = _build_merged_value_map(worksheet)
    year_markers = _extract_year_markers(worksheet, merged_values)
    events = []

    for start_col in MONTH_BLOCK_STARTS:
        month_num = _month_num_for_col(worksheet, merged_values, start_col)
        year = _year_for_column(year_markers, start_col)
        if month_num is None or year is None:
            continue

        active_label = None
        active_start = None
        active_end = None

        for row in range(GRID_START_ROW, GRID_END_ROW + 1):
            day_text = _clean_text(_cell_value(worksheet, merged_values, row, start_col))
            if not day_text.isdigit():
                continue

            try:
                current_date = datetime(year, month_num, int(day_text)).date()
            except ValueError:
                continue

            label = _pick_block_label(worksheet, merged_values, row, start_col)
            if label and _should_skip_grid_label(label):
                label = ""
            if label:
                label = _format_period_label(label, current_date)

            if not label:
                if active_label:
                    _append_range(events, active_label, active_start, active_end)
                    active_label = None
                    active_start = None
                    active_end = None
                continue

            if active_label == label and current_date == active_end + timedelta(days=1):
                active_end = current_date
                continue

            if active_label:
                _append_range(events, active_label, active_start, active_end)

            active_label = label
            active_start = current_date
            active_end = current_date

        if active_label:
            _append_range(events, active_label, active_start, active_end)

    return _merge_adjacent_ranges(events)


def _academic_start_year(worksheet, merged_values) -> int:
    for _, year in _extract_year_markers(worksheet, merged_values):
        return year
    return datetime.now().year


def _infer_year_from_month(month_num: int, academic_start_year: int) -> int:
    return academic_start_year if month_num >= 9 else academic_start_year + 1


def _expand_day_parts(days_raw: str):
    normalized = re.sub(r"\bet\b", ",", days_raw, flags=re.IGNORECASE)
    for part in [piece.strip() for piece in normalized.split(",") if piece.strip()]:
        if "-" in part:
            left, right = [piece.strip() for piece in part.split("-", 1)]
            if left.isdigit() and right.isdigit():
                yield int(left), int(right)
        elif part.isdigit():
            day = int(part)
            yield day, day


def _extract_text_events_from_sheet(worksheet):
    merged_values = _build_merged_value_map(worksheet)
    academic_start_year = _academic_start_year(worksheet, merged_values)
    events = []

    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            raw_value = _cell_value(worksheet, merged_values, row, col)
            if not isinstance(raw_value, str):
                continue

            text = _clean_text(raw_value)
            if not text:
                continue

            slash_match = SLASH_DATE_RE.search(text)
            if slash_match:
                label = text[slash_match.end():].strip(" -:;\t")
                if not label:
                    continue

                try:
                    current_date = datetime(
                        int(slash_match.group("year")),
                        int(slash_match.group("month")),
                        int(slash_match.group("day")),
                    ).date()
                except ValueError:
                    continue

                events.append(
                    {
                        "nom": label,
                        "date_debut": current_date,
                        "date_fin": current_date,
                        "type": _guess_type(label),
                    }
                )
                continue

            match = FRENCH_TEXT_DATE_RE.search(text)
            if not match:
                continue

            label = text[match.end():].strip(" -:;\t")
            if not label:
                continue

            month_key = _normalize_text(match.group("month")).lower()
            month_num = MONTHS.get(month_key)
            if month_num is None:
                continue

            explicit_year = match.group("year")
            year = int(explicit_year) if explicit_year else _infer_year_from_month(month_num, academic_start_year)

            for start_day, end_day in _expand_day_parts(match.group("days")):
                try:
                    start_date = datetime(year, month_num, start_day).date()
                    end_date = datetime(year, month_num, end_day).date()
                except ValueError:
                    continue

                events.append(
                    {
                        "nom": label,
                        "date_debut": start_date,
                        "date_fin": end_date,
                        "type": _guess_type(label),
                    }
                )

    return events


def _extract_footer_date_events_from_sheet(worksheet):
    events = []

    for row in range(FOOTER_START_ROW, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row, col).value
            if isinstance(value, datetime):
                current_date = value.date()
            elif isinstance(value, date):
                current_date = value
            else:
                continue

            label = ""
            for offset in range(1, 5):
                if col + offset > worksheet.max_column:
                    break
                candidate = _clean_text(worksheet.cell(row, col + offset).value)
                if candidate:
                    label = candidate
                    break

            if not label:
                continue

            events.append(
                {
                    "nom": label,
                    "date_debut": current_date,
                    "date_fin": current_date,
                    "type": _guess_type(label),
                }
            )

    return events


def extract_events_from_excel(excel_path: str):
    workbook = load_workbook(excel_path, data_only=True)
    all_events = []

    for worksheet in workbook.worksheets:
        all_events.extend(_extract_period_events_from_sheet(worksheet))
        all_events.extend(_extract_grid_events_from_sheet(worksheet))
        all_events.extend(_extract_text_events_from_sheet(worksheet))
        all_events.extend(_extract_footer_date_events_from_sheet(worksheet))

    dedup = {}
    for event in all_events:
        name = re.sub(r"\s+", " ", (event.get("nom") or "")).strip()
        if not name:
            continue

        key = (
            name.lower(),
            event["date_debut"].isoformat(),
            event["date_fin"].isoformat(),
            event["type"],
        )
        dedup[key] = {
            "nom": name,
            "date_debut": event["date_debut"],
            "date_fin": event["date_fin"],
            "type": event["type"],
        }

    return sorted(dedup.values(), key=lambda item: (item["date_debut"], item["date_fin"], item["nom"].lower()))


def import_calendar(excel_path: str, annee_id: int, clear_existing: bool = True, dry_run: bool = False):
    excel_path = os.path.abspath(excel_path)
    if not os.path.exists(excel_path):
        raise FileNotFoundError(excel_path)

    events = extract_events_from_excel(excel_path)

    print(f"Fichier: {excel_path}")
    print(f"Evenements detectes: {len(events)}")

    if dry_run:
        for event in sorted(events, key=lambda item: (item["date_debut"], item["nom"]))[:80]:
            print(f"- {event['date_debut']} -> {event['date_fin']} | {event['type']} | {event['nom']}")
        if len(events) > 80:
            print(f"... +{len(events) - 80} autres")
        return

    database_url = os.getenv(
        "DATABASE_URL",
        "postgresql://emploi_user:emploi_temps@127.0.0.1:5432/emploi_temps",
    )
    parsed = urlparse(database_url)

    conn = psycopg2.connect(
        host=parsed.hostname or "127.0.0.1",
        port=parsed.port or 5432,
        database=(parsed.path or "").lstrip("/") or "emploi_temps",
        user=parsed.username or "emploi_user",
        password=parsed.password or "emploi_temps",
    )

    try:
        with conn.cursor() as cur:
            if clear_existing:
                cur.execute("DELETE FROM vacances_jours_feries WHERE annee_id = %s", (annee_id,))
                deleted = cur.rowcount
                conn.commit()
                print(f"Suppression existant (annee_id={annee_id}): {deleted} lignes")

            cur.executemany(
                """
                INSERT INTO vacances_jours_feries (nom, date_debut, date_fin, type, annee_id)
                VALUES (%s, %s, %s, %s, %s)
                """,
                [(event["nom"], event["date_debut"], event["date_fin"], event["type"], annee_id) for event in events],
            )
            conn.commit()

        print(f"Import termine: {len(events)} lignes inserees (annee_id={annee_id})")
    finally:
        conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--file",
        default=os.path.join("public", "excel_files", "Calendrier universitaire 2024-2025 Modifié.xlsx"),
    )
    parser.add_argument("--annee-id", type=int, default=1)
    parser.add_argument("--no-clear", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    import_calendar(
        excel_path=args.file,
        annee_id=args.annee_id,
        clear_existing=not args.no_clear,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
